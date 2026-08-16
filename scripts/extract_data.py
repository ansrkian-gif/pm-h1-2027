"""
Extract NTE work orders from PM Excel (Complete Operator contains 'NTE')
and write dashboard-data.js for the dashboard.

Also builds site integrity checks across the FULL report:
- General sites need BOTH Passive General + Active General
- Small Cell sites should be a single Site ID entry
"""
from __future__ import annotations

import json
import os
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl first: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DASH = ROOT / "dashboard"
OUT = DASH / "dashboard-data.js"

CYCLE_START = "2026-08-01"
CYCLE_END = "2027-01-31"
DAILY_TEAM_TARGET = 21
PER_FME_DAILY_TARGET = 3
ACTIVE_FME_COUNT = 7

PASSIVE_GENERAL = "OGK Passive General"
ACTIVE_GENERAL = "OGK Active General"
SMALL_CELL = "OGK Active Small Cell /Book RRU/Easy Macro"

# Nabi replaced Mohd — count both under one FME seat
FME_ALIASES = {
    "Mohd": "Mohd+Nabi",
    "Nabi": "Mohd+Nabi",
}


def find_xlsx() -> Path:
    env = os.environ.get("EXTRACT_PATH")
    if env and Path(env).exists():
        return Path(env)
    cloud = ROOT / "cloud" / "extract.xlsx"
    if cloud.exists():
        return cloud
    files = [
        p
        for p in ROOT.glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not files:
        raise SystemExit(f"No .xlsx file found in {ROOT}")
    return max(files, key=lambda p: p.stat().st_mtime)


def short_name(operator: str) -> str:
    parts = operator.split("_")
    raw = parts[-1] if parts else operator
    return FME_ALIASES.get(raw, raw)


def is_small_cell(sub: str) -> bool:
    return "Small Cell" in (sub or "")


def resolve_task_id_col(headers: dict) -> int:
    for name in ("Task Id", "Task ID", "TaskId", "s"):
        if name in headers:
            return headers[name]
    # Fallback: first column
    return 1


def load_exempt_task_ids() -> set[str]:
    path = ROOT / "exempt_task_ids.txt"
    if not path.exists():
        return set()
    ids: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        ids.add(line)
    return ids


def build_site_checks(ws, headers: dict, exempt_ids: set[str]) -> dict:
    """Analyze ALL rows (not only NTE) for Site ID / subcategory integrity.

    Exempt Task IDs are excluded from pairing / Small Cell rules and summarized separately.
    """
    task_col = resolve_task_id_col(headers)
    site_subs: dict[str, Counter] = defaultdict(Counter)
    sub_total = Counter()
    exempt_rows = []
    exempt_sites: Counter = Counter()
    listed = set(exempt_ids)
    found_exempt: set[str] = set()

    for r in range(2, ws.max_row + 1):
        task_id = str(ws.cell(r, task_col).value or "").strip()
        site_raw = ws.cell(r, headers["Site ID"]).value if "Site ID" in headers else None
        sub_raw = ws.cell(r, headers["Task Subcategory"]).value if "Task Subcategory" in headers else None
        site = str(site_raw).strip() if site_raw is not None else ""
        sub = str(sub_raw).strip() if sub_raw is not None else ""
        op = ws.cell(r, headers["Complete Operator"]).value if "Complete Operator" in headers else None

        if task_id in exempt_ids:
            found_exempt.add(task_id)
            if site:
                exempt_sites[site] += 1
            exempt_rows.append(
                {
                    "taskId": task_id,
                    "siteId": site,
                    "taskSubcategory": sub,
                    "completeOperator": str(op).strip() if op else "",
                }
            )
            continue  # exempt from integrity rules

        if not site:
            continue
        site_subs[site][sub] += 1
        if sub:
            sub_total[sub] += 1

    pair_complete = []
    incomplete = []
    small_sites = []
    small_duplicates = []
    other_sites = []

    for site, counts in sorted(site_subs.items(), key=lambda kv: kv[0]):
        keys = set(counts.keys())
        has_p = PASSIVE_GENERAL in keys
        has_a = ACTIVE_GENERAL in keys
        small_count = sum(v for k, v in counts.items() if is_small_cell(k))
        has_s = small_count > 0

        if has_s and not has_p and not has_a:
            small_sites.append(
                {
                    "siteId": site,
                    "woCount": small_count,
                    "subcategory": SMALL_CELL,
                }
            )
            if small_count > 1:
                small_duplicates.append({"siteId": site, "woCount": small_count})
            continue

        if has_p and has_a:
            pair_complete.append(site)
            continue

        if has_p or has_a:
            missing = []
            present = []
            if has_p:
                present.append(PASSIVE_GENERAL)
            else:
                missing.append(PASSIVE_GENERAL)
            if has_a:
                present.append(ACTIVE_GENERAL)
            else:
                missing.append(ACTIVE_GENERAL)
            incomplete.append(
                {
                    "siteId": site,
                    "present": present,
                    "missing": missing,
                    "allSubs": sorted([k for k in keys if k]),
                }
            )
            continue

        other_sites.append({"siteId": site, "subs": sorted([k for k in keys if k])})

    missing_passive = sum(1 for x in incomplete if PASSIVE_GENERAL in x["missing"])
    missing_active = sum(1 for x in incomplete if ACTIVE_GENERAL in x["missing"])
    not_in_file = sorted(listed - found_exempt)

    return {
        "rules": {
            "generalPair": [PASSIVE_GENERAL, ACTIVE_GENERAL],
            "generalExpectedEntries": 2,
            "smallCell": SMALL_CELL,
            "smallCellExpectedEntries": 1,
        },
        "subcategoryTotals": dict(sub_total.most_common()),
        "smallCell": {
            "label": SMALL_CELL,
            "woCount": int(sub_total.get(SMALL_CELL, 0))
            or sum(v for k, v in sub_total.items() if is_small_cell(k)),
            "siteCount": len(small_sites),
            "sites": small_sites,
            "duplicateSites": small_duplicates,
            "duplicateCount": len(small_duplicates),
        },
        "generalPairs": {
            "completeSiteCount": len(pair_complete),
            "incompleteSiteCount": len(incomplete),
            "missingPassiveCount": missing_passive,
            "missingActiveCount": missing_active,
            "incompleteSites": incomplete,
        },
        "exemptions": {
            "listedWoCount": len(listed),
            "matchedWoCount": len(found_exempt),
            "missingWoCount": len(not_in_file),
            "missingTaskIds": not_in_file,
            "commonSiteIdCount": len(exempt_sites),
            "siteIds": [
                {"siteId": site, "woCount": count}
                for site, count in sorted(exempt_sites.items(), key=lambda kv: (-kv[1], kv[0]))
            ],
            "workOrders": exempt_rows,
        },
        "otherSiteCount": len(other_sites),
        "totalSites": len(site_subs),
    }


def main() -> None:
    xlsx = find_xlsx()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    task_col = resolve_task_id_col(headers)
    required = ["Complete Operator", "Complete Time", "Task Status", "Site ID", "Task Subcategory"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"Missing columns: {missing}")

    exempt_ids = load_exempt_task_ids()

    records = []
    for r in range(2, ws.max_row + 1):
        op = ws.cell(r, headers["Complete Operator"]).value
        if not op or "NTE" not in str(op).upper():
            continue

        task_id = str(ws.cell(r, task_col).value or "").strip()
        # Exempt WOs are excluded from NTE performance totals
        if task_id in exempt_ids:
            continue

        ct = ws.cell(r, headers["Complete Time"]).value
        if isinstance(ct, datetime):
            complete_iso = ct.strftime("%Y-%m-%d %H:%M:%S")
            complete_date = ct.strftime("%Y-%m-%d")
        elif ct:
            complete_iso = str(ct)
            complete_date = str(ct)[:10]
        else:
            complete_iso = None
            complete_date = None

        name = str(op).strip()
        sub = ws.cell(r, headers["Task Subcategory"]).value
        records.append(
            {
                "taskId": task_id,
                "title": ws.cell(r, headers["Title"]).value if "Title" in headers else None,
                "siteId": str(ws.cell(r, headers["Site ID"]).value)
                if ws.cell(r, headers["Site ID"]).value is not None
                else "",
                "taskSubcategory": str(sub).strip() if sub is not None else "",
                "status": ws.cell(r, headers["Task Status"]).value,
                "completeOperator": name,
                "fmeShort": short_name(name),
                "completeTime": complete_iso,
                "completeDate": complete_date,
                "assignTo": ws.cell(r, headers["Assign To FME Full Name"]).value
                if "Assign To FME Full Name" in headers
                else None,
            }
        )

    site_checks = build_site_checks(ws, headers, exempt_ids)
    fme_counts = Counter(rec["fmeShort"] for rec in records)
    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sourceFile": xlsx.name,
        "cycle": {"start": CYCLE_START, "end": CYCLE_END},
        "targets": {
            "dailyTeam": DAILY_TEAM_TARGET,
            "perFmeDaily": PER_FME_DAILY_TARGET,
            "activeFmeCount": ACTIVE_FME_COUNT,
            "excludeWeekday": 5,  # Friday (JS getDay(): 0=Sun ... 5=Fri)
        },
        "fmeTotals": dict(sorted(fme_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
        "siteChecks": site_checks,
        "workOrders": records,
    }

    OUT.write_text(
        "window.DASHBOARD_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT.name} from {xlsx.name}")
    print(f"NTE WOs (after exemptions): {len(records)} | FMEs: {len(fme_counts)}")
    for name, count in fme_counts.most_common():
        print(f"  {count:3d}  {name}")
    ex = site_checks["exemptions"]
    print(
        f"Exempted: {ex['matchedWoCount']}/{ex['listedWoCount']} WOs | "
        f"common Site IDs: {ex['commonSiteIdCount']}"
    )
    gp = site_checks["generalPairs"]
    sc = site_checks["smallCell"]
    print(
        f"Site checks: pair OK {gp['completeSiteCount']} | incomplete {gp['incompleteSiteCount']} "
        f"| Small Cell sites {sc['siteCount']} (WOs {sc['woCount']})"
    )
    if gp["incompleteSiteCount"]:
        print("  Incomplete Active/Passive sites:")
        for row in gp["incompleteSites"][:15]:
            print(f"    {row['siteId']}: missing {', '.join(row['missing'])}")


if __name__ == "__main__":
    main()
